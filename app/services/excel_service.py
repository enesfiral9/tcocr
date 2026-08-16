from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from app.schemas import ExportRecord


HEADERS = ["Sıra No", "T.C. Kimlik No", "Ad", "Soyad", "Doğum Tarihi", "Seri No",
           "Son Geçerlilik", "Cinsiyet", "Uyruk", "Anne Adı", "Baba Adı", "Veren Makam",
           "OCR Güven Oranı", "Kontrol Durumu"]


def create_excel(records: list[ExportRecord]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "OCR Sonuçları"
    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="16324F")
    for index, record in enumerate(records, 1):
        sheet.append([index, record.tc_no, record.name, record.surname, record.birth_date, record.serial_no,
                      record.expiry_date, record.gender, record.nationality, record.mother_name,
                      record.father_name, record.issuing_authority, record.confidence,
                      "Kontrol Gerekli" if record.requires_review else "Başarılı"])
        sheet.cell(index + 1, 13).number_format = "0%"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:N{max(1, sheet.max_row)}"
    for index, width in enumerate([10, 18, 22, 22, 18, 18, 18, 12, 14, 22, 22, 24, 20, 22], 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
